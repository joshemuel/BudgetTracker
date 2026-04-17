"""One-off importer: BudgetTrackerPrototype.xlsx → Postgres.

Reads the four-sheet legacy workbook and loads it into the new schema for a
single user (``josia`` by default). Idempotent: re-running deletes that user's
existing data and re-imports from the xlsx. Per-source ``starting_balance`` is
back-computed so that ``starting + sum(income) - sum(expense)`` equals the
current balance shown in the sheet.
"""

from __future__ import annotations

import argparse
import os
from collections import defaultdict
from datetime import date, datetime, time, timezone
from decimal import Decimal
from pathlib import Path
from uuid import uuid4

from openpyxl import load_workbook
from sqlalchemy import delete
from sqlalchemy.orm import Session

from app.db.models import (
    DEFAULT_CATEGORIES,
    Budget,
    Category,
    Source,
    Transaction,
    User,
)
from app.db.session import SessionLocal
from app.services.auth import hash_password

DEFAULT_XLSX = Path("/seed/BudgetTrackerPrototype.xlsx")
DEFAULT_USERNAME = "josia"


def _decimal(value) -> Decimal:
    if value is None:
        return Decimal("0")
    if isinstance(value, Decimal):
        return value
    # openpyxl returns float/int for numeric cells
    return Decimal(str(value))


def _combine_dt(d, t) -> datetime:
    if isinstance(d, datetime):
        d = d.date()
    if t is None:
        t = time(0, 0, 0)
    if isinstance(t, datetime):
        t = t.time()
    return datetime.combine(d, t, tzinfo=timezone.utc)


def _read_sources(ws) -> list[dict]:
    """Read the 'Source' section of the BudgetTracker sheet.

    The section starts at row 1 ('Source' header), row 2 is column headers,
    rows 3+ are data until a blank Method cell.
    """
    sources = []
    for row in range(3, ws.max_row + 1):
        method = ws.cell(row, 1).value
        if method is None or str(method).strip() == "":
            break
        if str(method).strip() == "Limits":  # section divider reached
            break
        amount = ws.cell(row, 2).value
        sources.append({"name": str(method).strip(), "current_amount": _decimal(amount)})
    return sources


def _read_limits(ws) -> list[dict]:
    """Read the 'Limits' section of the BudgetTracker sheet."""
    # Find the 'Limits' header row, then data starts two rows below.
    header_row = None
    for row in range(1, ws.max_row + 1):
        if ws.cell(row, 1).value == "Limits":
            header_row = row
            break
    if header_row is None:
        return []

    limits = []
    for row in range(header_row + 2, ws.max_row + 1):
        cat = ws.cell(row, 1).value
        if cat is None or str(cat).strip() == "":
            break
        if str(cat).strip() in {"Credit Card", "Custom Categories"}:
            break
        amount = ws.cell(row, 2).value
        limits.append({"category": str(cat).strip(), "monthly_limit": _decimal(amount)})
    return limits


def _read_transactions(ws) -> list[dict]:
    """Read RawBudget transaction log (header in row 1, data rows 2..N)."""
    txns = []
    for row in range(2, ws.max_row + 1):
        date_cell = ws.cell(row, 1).value
        if date_cell is None:
            continue
        time_cell = ws.cell(row, 2).value
        type_cell = ws.cell(row, 3).value
        category_cell = ws.cell(row, 4).value
        amount_cell = ws.cell(row, 5).value
        source_cell = ws.cell(row, 6).value
        desc_cell = ws.cell(row, 7).value
        if type_cell is None or category_cell is None:
            continue
        # Legacy sheet had 56 rows with blank Source; Apps Script defaulted to 'BCA'.
        source_name = str(source_cell).strip() if source_cell is not None else "BCA"
        txns.append(
            {
                "occurred_at": _combine_dt(date_cell, time_cell),
                "type": str(type_cell).strip().lower(),  # 'expense' | 'income'
                "category": str(category_cell).strip(),
                "amount": _decimal(amount_cell),
                "source": source_name,
                "description": (str(desc_cell).strip() if desc_cell is not None else None),
            }
        )
    return txns


def _detect_transfers(txns: list[dict]) -> None:
    """Mark paired Top-up Expense/Income rows with a shared transfer_group_id.

    Pairs are matched on (occurred_at.date, amount) with one expense + one
    income, both category 'Top-up'. Order-based matching breaks ties for
    multiple same-day same-amount transfers.
    """
    by_key: dict[tuple[date, Decimal], dict[str, list[dict]]] = defaultdict(
        lambda: {"expense": [], "income": []}
    )
    for t in txns:
        if t["category"] == "Top-up":
            key = (t["occurred_at"].date(), t["amount"])
            by_key[key][t["type"]].append(t)
    for (_day, _amt), sides in by_key.items():
        for exp, inc in zip(sides["expense"], sides["income"]):
            group_id = uuid4()
            exp["transfer_group_id"] = group_id
            inc["transfer_group_id"] = group_id


def run_import(
    db: Session,
    xlsx_path: Path,
    username: str,
    password: str,
    telegram_chat_id: str | None,
    preserve_transactions: bool,
) -> dict:
    wb = load_workbook(xlsx_path, data_only=True)
    src_rows = _read_sources(wb["BudgetTracker"])
    budget_rows = _read_limits(wb["BudgetTracker"])
    txn_rows = _read_transactions(wb["RawBudget"])
    _detect_transfers(txn_rows)

    user = db.query(User).filter_by(username=username).one_or_none()
    if user is None:
        user = User(
            username=username,
            password_hash=hash_password(password),
            telegram_chat_id=telegram_chat_id or None,
        )
        db.add(user)
        db.flush()
    else:
        # Refresh password + chat id on re-import so env changes take effect
        user.password_hash = hash_password(password)
        user.telegram_chat_id = telegram_chat_id or None
        if not preserve_transactions:
            db.execute(delete(Transaction).where(Transaction.user_id == user.id))
            db.execute(delete(Budget).where(Budget.user_id == user.id))
            db.execute(delete(Source).where(Source.user_id == user.id))
            db.execute(delete(Category).where(Category.user_id == user.id))
            db.flush()

    cats_by_name: dict[str, Category] = {}
    for name in DEFAULT_CATEGORIES:
        cat = Category(user_id=user.id, name=name, is_default=True)
        db.add(cat)
        cats_by_name[name] = cat
    # any category referenced by transactions/budgets that isn't a default
    extra_names = {t["category"] for t in txn_rows} | {b["category"] for b in budget_rows}
    extra_names -= set(cats_by_name)
    for name in sorted(extra_names):
        cat = Category(user_id=user.id, name=name, is_default=False)
        db.add(cat)
        cats_by_name[name] = cat
    db.flush()

    sources_by_name: dict[str, Source] = {}
    for s in src_rows:
        obj = Source(
            user_id=user.id,
            name=s["name"],
            starting_balance=Decimal("0"),  # computed after transactions insert
            currency="IDR",
            is_credit_card=("credit card" in s["name"].lower()),
            active=True,
        )
        db.add(obj)
        sources_by_name[s["name"]] = obj
    # any source referenced by transactions that isn't in the Source section
    for t in txn_rows:
        if t["source"] not in sources_by_name:
            obj = Source(
                user_id=user.id,
                name=t["source"],
                starting_balance=Decimal("0"),
                currency="IDR",
                is_credit_card=("credit card" in t["source"].lower()),
                active=True,
            )
            db.add(obj)
            sources_by_name[t["source"]] = obj
    db.flush()

    for b in budget_rows:
        cat = cats_by_name[b["category"]]
        db.add(Budget(user_id=user.id, category_id=cat.id, monthly_limit=b["monthly_limit"]))

    # Bulk insert transactions
    to_insert = []
    for t in txn_rows:
        cat = cats_by_name[t["category"]]
        src = sources_by_name[t["source"]]
        to_insert.append(
            Transaction(
                user_id=user.id,
                occurred_at=t["occurred_at"],
                type=t["type"],
                category_id=cat.id,
                amount=t["amount"],
                source_id=src.id,
                description=t["description"],
                transfer_group_id=t.get("transfer_group_id"),
            )
        )
    db.add_all(to_insert)
    db.flush()

    # Back-compute starting_balance for each source so final balance = sheet
    sum_by_source: dict[str, dict[str, Decimal]] = defaultdict(
        lambda: {"income": Decimal("0"), "expense": Decimal("0")}
    )
    for t in txn_rows:
        sum_by_source[t["source"]][t["type"]] += t["amount"]
    current_by_sheet = {s["name"]: s["current_amount"] for s in src_rows}
    for name, src in sources_by_name.items():
        current = current_by_sheet.get(name, Decimal("0"))
        inc = sum_by_source[name]["income"]
        exp = sum_by_source[name]["expense"]
        src.starting_balance = current - inc + exp

    db.commit()

    return {
        "user_id": user.id,
        "sources": len(sources_by_name),
        "categories": len(cats_by_name),
        "budgets": len(budget_rows),
        "transactions": len(to_insert),
    }


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX)
    parser.add_argument("--username", default=DEFAULT_USERNAME)
    parser.add_argument(
        "--password",
        default=os.environ.get("SEED_PASSWORD", "changeme"),
        help="Defaults to SEED_PASSWORD env var, then 'changeme'.",
    )
    parser.add_argument(
        "--telegram-chat-id",
        default=os.environ.get("TELEGRAM_CHAT_ID", "") or None,
    )
    parser.add_argument("--preserve-transactions", action="store_true")
    args = parser.parse_args()

    with SessionLocal() as db:
        stats = run_import(
            db,
            xlsx_path=args.xlsx,
            username=args.username,
            password=args.password,
            telegram_chat_id=args.telegram_chat_id,
            preserve_transactions=args.preserve_transactions,
        )
    print(stats)


if __name__ == "__main__":
    main()
